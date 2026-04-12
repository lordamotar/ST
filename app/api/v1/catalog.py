import os
import uuid
import re
from pathlib import Path

from fastapi import APIRouter, Depends, Query, HTTPException, Header, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from typing import List, Optional

from app.core.database import get_db
from app.core.config import settings
from app.models.product import Product, Category
from app.models.user import User
from app.schemas.product import (
    ProductResponse, CategoryResponse, ProductCreate, ProductUpdate,
    CategoryCreate, CategoryUpdate,
    BulkImportResponse, BulkImportError,
)
from app.core.dependencies import check_admin, require_roles, get_current_user
from loguru import logger

router = APIRouter()

# --- Директория для загрузок ---
UPLOAD_DIR = Path("uploads/products")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/webp", "image/gif"}
MAX_IMAGE_SIZE = 5 * 1024 * 1024  # 5 MB


# Код удален: использование новой системы JWT


# =====================================================================
# PUBLIC ENDPOINTS
# =====================================================================

@router.get("/categories", response_model=List[CategoryResponse])
async def get_categories(db: AsyncSession = Depends(get_db)):
    """Все категории мебели."""
    result = await db.execute(select(Category))
    return result.scalars().all()


@router.get("/products", response_model=List[ProductResponse])
async def get_products(
    category_slug: Optional[str] = None,
    material: Optional[str] = None,
    color: Optional[str] = None,
    slug: Optional[str] = None,
    q: Optional[str] = None,
    is_bestseller: Optional[bool] = None,
    is_active: Optional[bool] = None,
    all_products: bool = Query(False),
    sort: Optional[str] = None,
    db: AsyncSession = Depends(get_db)
):
    query = select(Product).join(Category).options(selectinload(Product.category))

    if is_active is not None:
        query = query.where(Product.is_active == is_active)
    elif not all_products:
        query = query.where(Product.is_active == True)

    if category_slug:
        query = query.where(Category.slug == category_slug)
    if material:
        query = query.where(Product.material == material)
    if color:
        query = query.where(Product.color == color)
    if slug:
        query = query.where(Product.slug == slug)
    if is_bestseller is not None:
        query = query.where(Product.is_bestseller == is_bestseller)
    if q:
        query = query.where(
            Product.name.ilike(f"%{q}%") |
            func.coalesce(Product.description, "").ilike(f"%{q}%")
        )

    # Sorting
    if sort == "price_asc":
        query = query.order_by(Product.new_price.asc())
    elif sort == "price_desc":
        query = query.order_by(Product.new_price.desc())
    elif sort == "newest":
        query = query.order_by(Product.id.desc())
    else:
        # Default sort
        query = query.order_by(Product.id.desc())

    result = await db.execute(query)
    return result.scalars().all()


@router.get("/products/template")
async def download_excel_template(
    admin: User = Depends(require_roles(["admin", "manager"]))
):
    """
    Скачать шаблон Excel для массового импорта товаров.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"

    # Заголовки
    headers = [
        "name", "new_price", "old_price", "category_id", "slug",
        "material", "color", "description", "image_url", "is_active", "availability_status",
        "show_timer", "promo_start", "promo_end",
        "Размеры", "Материал ножек (опор)", "Материал столешницы", "Толщина столешницы",
        "Просвет от пола", "Максимальная нагрузка", "Регулировка опор", "Цвет столешницы",
        "Подпятники", "Гарантия", "Вариант доставки", "Опоры", "Страна", "Серия"
    ]
    header_labels = [
        "Название товара *", "Новая цена (₸) *", "Старая цена (₸)", "ID Категории *", "Slug (URL)",
        "Материал", "Цвет", "Описание товара", "Ссылка на фото", "Активен (TRUE/FALSE)", "Наличие (в наличии/под заказ)",
        "Показать таймер (TRUE/FALSE)", "Дата начала акции (ГГГГ-ММ-ДД)", "Дата окончания акции (ГГГГ-ММ-ДД)",
        "Размеры", "Материал ножек (опор)", "Материал столешницы", "Толщина столешницы",
        "Просвет от пола", "Максимальная нагрузка", "Регулировка опор", "Цвет столешницы",
        "Подпятники", "Гарантия", "Вариант доставки", "Опоры", "Страна", "Серия"
    ]

    # Стили заголовков
    header_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Строка 1: Машинные заголовки (для парсинга)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Строка 2: Описание полей
    desc_font = Font(name="Arial", size=9, italic=True, color="888888")
    for col_idx, label in enumerate(header_labels, 1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = desc_font
        cell.alignment = Alignment(horizontal="center")

    # Примеры данных (строки 3-5)
    examples = [
        ["Стул 'Осло' из дуба", 125000, 150000, 1, "oslo-oak-chair", "oak", "natural",
         "Эргономичный стул из массива дуба.",
         "https://example.com/photos/oslo-chair.jpg", True, "в наличии",
         "Длина 120 смхШирина 60 смхВысота 75 см", "сталь", "МДФ 15 мм", "1.5 см", "12 см", "15 кг", "нет", "Дуб Натуральный", "пластиковые", "12 месяцев", "в разобранном виде", "Черный", "Китай", "Рико"]
    ]

    data_font = Font(name="Arial", size=10)
    for row_idx, row_data in enumerate(examples, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Ширина колонок
    col_widths = [30, 12, 14, 25, 15, 12, 50, 40, 18, 25] + [15]*14
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # Высота строк с примерами
    for row_idx in range(3, 6):
        ws.row_dimensions[row_idx].height = 45

    # Лист 2: Справочник категорий
    ws2 = wb.create_sheet(title="Категории (справочник)")
    ws2.cell(row=1, column=1, value="ID").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value="Название").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.cell(row=1, column=3, value="Slug").font = header_font
    ws2.cell(row=1, column=3).fill = header_fill

    cat_examples = [
        [1, "Стулья", "chairs"],
        [2, "Диваны", "sofas"],
        [3, "Столы", "tables"],
        [4, "Шкафы", "cupboards"],
    ]
    for row_idx, cat in enumerate(cat_examples, 2):
        for col_idx, value in enumerate(cat, 1):
            ws2.cell(row=row_idx, column=col_idx, value=value).font = data_font
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 20

    # Сохраняем в буфер
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    wb.close()

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=products_import_template.xlsx"}
    )


@router.get("/products/{slug}", response_model=ProductResponse)
async def get_product_by_slug(slug: str, db: AsyncSession = Depends(get_db)):
    """Получить один товар по его slug."""
    result = await db.execute(
        select(Product).where(Product.slug == slug).options(selectinload(Product.category))
    )
    db_product = result.scalar_one_or_none()
    if not db_product:
        raise HTTPException(status_code=404, detail="Товар не найден")
    
    # Скрываем неактивные товары для публичного доступа
    if not db_product.is_active:
        raise HTTPException(status_code=404, detail="Товар временно недоступен")
        
    return db_product


@router.get("/categories/{slug}", response_model=CategoryResponse)
async def get_category_by_slug(slug: str, db: AsyncSession = Depends(get_db)):
    """Получить одну категорию по её slug."""
    result = await db.execute(select(Category).where(Category.slug == slug))
    db_cat = result.scalar_one_or_none()
    if not db_cat:
        raise HTTPException(status_code=404, detail="Категория не найдена")
    return db_cat


# =====================================================================
# ADMIN ENDPOINTS — PRODUCT CRUD
# =====================================================================

@router.post("/products", response_model=ProductResponse, status_code=201)
async def create_product(
    product_in: ProductCreate,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_roles(["admin", "manager"]))
):
    """Создать новый товар (Админ или Менеджер)."""
    db_product = Product(**product_in.model_dump())
    db.add(db_product)
    await db.commit()
    await db.refresh(db_product)
    logger.info(f"New product created: {db_product.name} (ID: {db_product.id}, Price: {db_product.new_price} ₸) by {admin.username}")

    # Подгружаем связанную категорию для ответа
    result = await db.execute(
        select(Product).where(Product.id == db_product.id).options(selectinload(Product.category))
    )
    return result.scalar_one()


@router.patch("/products/{product_id}", response_model=ProductResponse)
async def update_product(
    product_id: int,
    product_in: ProductUpdate,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_roles(["admin", "manager"]))
):
    """Обновить товар (только для администратора)."""
    result = await db.execute(
        select(Product).where(Product.id == product_id).options(selectinload(Product.category))
    )
    db_product = result.scalar_one_or_none()
    if not db_product:
        raise HTTPException(status_code=404, detail="Product not found")

    # Обновляем только переданные поля
    update_data = product_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_product, field, value)

    await db.commit()
    await db.refresh(db_product)

    result = await db.execute(
        select(Product).where(Product.id == product_id).options(selectinload(Product.category))
    )
    return result.scalar_one()


@router.delete("/products/{product_id}", status_code=204)
async def delete_product(
    product_id: int,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_roles(["admin"]))
):
    """Удалить товар (только для администратора)."""
    result = await db.execute(select(Product).where(Product.id == product_id))
    db_product = result.scalar_one_or_none()
    if not db_product:
        raise HTTPException(status_code=404, detail="Product not found")

    await db.delete(db_product)
    await db.commit()
    logger.info(f"Product deleted: {db_product.name} (ID: {db_product.id})")


# =====================================================================
# ADMIN ENDPOINT — QUICK TOGGLE ACTIVE/INACTIVE
# =====================================================================

@router.patch("/products/{product_id}/toggle", response_model=ProductResponse)
async def toggle_product_status(
    product_id: int,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_roles(["admin", "manager"]))
):
    """Быстрое переключение статуса товара (активен/неактивен)."""
    result = await db.execute(
        select(Product).where(Product.id == product_id).options(selectinload(Product.category))
    )
    db_product = result.scalar_one_or_none()
    if not db_product:
        raise HTTPException(status_code=404, detail="Product not found")

    db_product.is_active = not db_product.is_active
    await db.commit()
    await db.refresh(db_product)
    logger.info(f"Product status toggled: {db_product.name} (ID: {db_product.id}, Active: {db_product.is_active})")

    result = await db.execute(
        select(Product).where(Product.id == product_id).options(selectinload(Product.category))
    )
    return result.scalar_one()


# =====================================================================
# ADMIN ENDPOINT — IMAGE UPLOAD
# =====================================================================

@router.post("/upload-image")
async def upload_product_image(
    file: UploadFile = File(...),
    admin: User = Depends(require_roles(["admin", "manager"]))
):
    """Загрузить изображение товара. Возвращает URL загруженного файла."""

    # Проверка типа файла
    if file.content_type not in ALLOWED_IMAGE_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"Неподдерживаемый формат: {file.content_type}. "
                   f"Допустимо: JPEG, PNG, WebP, GIF."
        )

    # Проверка размера
    contents = await file.read()
    if len(contents) > MAX_IMAGE_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"Файл слишком большой ({len(contents) // 1024 // 1024} MB). "
                   f"Максимум: {MAX_IMAGE_SIZE // 1024 // 1024} MB."
        )

    # Генерация уникального имени
    ext = Path(file.filename).suffix.lower() if file.filename else ".jpg"
    if ext not in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
        ext = ".jpg"
    unique_name = f"{uuid.uuid4().hex}{ext}"
    file_path = UPLOAD_DIR / unique_name

    # Сохранение файла
    with open(file_path, "wb") as f:
        f.write(contents)

    return {"url": f"/uploads/products/{unique_name}"}


# =====================================================================
# ADMIN ENDPOINT — BULK IMPORT FROM EXCEL
# =====================================================================

def _slugify(text: str) -> str:
    """Генерация slug из названия товара."""
    # Транслитерация кириллицы
    translit_map = {
        'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'yo',
        'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
        'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
        'ф': 'f', 'х': 'kh', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'shch',
        'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
    }
    result = []
    for char in text.lower():
        if char in translit_map:
            result.append(translit_map[char])
        elif char.isascii() and (char.isalnum() or char in (' ', '-')):
            result.append(char)
        else:
            result.append(' ')
    slug = '-'.join(''.join(result).split())
    return slug or f"product-{uuid.uuid4().hex[:8]}"


@router.post("/products/bulk-import", response_model=BulkImportResponse)
async def bulk_import_products(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_roles(["admin"]))
):
    """
    Массовый импорт товаров из Excel (.xlsx).

    Ожидаемые столбцы:
    | name | price | category_id | slug | material | color | description | image_url | is_active |

    Любые другие столбцы (например "Страна", "Гарантия") будут автоматически добавлены как характеристики.

    - `name`, `price`, `category_id` — обязательные.
    - `slug` — автоматически генерируется из name, если не указан.
    - `image_url` — ссылка на изображение (полный URL или путь /uploads/...).
    - `is_active` — по умолчанию TRUE.
    """
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Файл должен быть в формате .xlsx")

    try:
        from openpyxl import load_workbook
        import io

        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), read_only=True)
        ws = wb.active

        if not ws:
            raise HTTPException(status_code=400, detail="Файл пуст или не содержит активного листа")

        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if len(rows) < 2:
            raise HTTPException(status_code=400, detail="Файл пуст (нужна строка заголовков + данные)")

        # Парсим заголовки
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        original_headers = [str(h).strip() if h else "" for h in rows[0]]

        # Маппинг колонок
        col_map = {}
        dynamic_cols = {}
        expected = [
            "name", "new_price", "old_price", "category_id", "slug", "material", "color", "description", "image_url", "is_active", "availability_status",
            "show_timer", "promo_start", "promo_end",
            "dimensions", "legs_material", "tabletop_material", "tabletop_thickness", "floor_clearance", "max_load",
            "legs_adjustment", "tabletop_color", "footings", "warranty", "delivery_format", "supports", "country", "series"
        ]
        
        # Переводы для удобства пользователей
        aliases = {
            "новая цена": "new_price",
            "старая цена": "old_price",
            "цена": "new_price",
            "размеры": "dimensions",
            "материал ножек (опор)": "legs_material",
            "материал столешницы": "tabletop_material",
            "толщина столешницы": "tabletop_thickness",
            "просвет от пола": "floor_clearance",
            "максимальная нагрузка": "max_load",
            "регулировка опор": "legs_adjustment",
            "цвет столешницы": "tabletop_color",
            "подпятники": "footings",
            "гарантия": "warranty",
            "вариант доставки": "delivery_format",
            "опоры": "supports",
            "страна": "country",
            "серия": "series",
            "наличие": "availability_status",
            "статус наличия": "availability_status",
            "показать таймер": "show_timer",
            "дата начала акции": "promo_start",
            "дата окончания акции": "promo_end"
        }

        for idx, (h_lower, h_orig) in enumerate(zip(headers, original_headers)):
            mapped_h = aliases.get(h_lower, h_lower)
            if mapped_h in expected:
                col_map[mapped_h] = idx
            elif h_lower and not h_lower.startswith("unnamed"):
                dynamic_cols[h_orig] = idx

        if "name" not in col_map or "new_price" not in col_map:
            raise HTTPException(
                status_code=400,
                detail="В файле должны быть столбцы 'name' и 'новая цена'. "
                       f"Найденные заголовки: {headers}"
            )

        # Загружаем существующие slugs для проверки дубликатов
        existing_slugs_result = await db.execute(select(Product.slug))
        existing_slugs = set(existing_slugs_result.scalars().all())

        created = 0
        skipped = 0
        errors: List[BulkImportError] = []

        for row_idx, row in enumerate(rows[1:], start=2):
            try:
                name = str(row[col_map["name"]]).strip() if row[col_map["name"]] else None
                if not name:
                    skipped += 1
                    continue

                new_price = float(row[col_map["new_price"]]) if row[col_map["new_price"]] else None
                if new_price is None or new_price <= 0:
                    errors.append(BulkImportError(row=row_idx, error=f"Некорректная цена: {row[col_map['new_price']]}"))
                    continue

                old_price = float(row[col_map["old_price"]]) if col_map.get("old_price") is not None and row[col_map["old_price"]] else None

                category_id = int(row[col_map.get("category_id", -1)]) if col_map.get("category_id") is not None and row[col_map["category_id"]] else None
                if not category_id:
                    errors.append(BulkImportError(row=row_idx, error="Не указан category_id"))
                    continue

                # Slug
                slug_val = row[col_map["slug"]] if col_map.get("slug") is not None else None
                slug = str(slug_val).strip() if slug_val is not None else None
                if slug and slug.lower() == "none":
                    slug = None
                if not slug:
                    slug = _slugify(name)

                # Проверка уникальности slug
                base_slug = slug
                counter = 1
                while slug in existing_slugs:
                    slug = f"{base_slug}-{counter}"
                    counter += 1

                existing_slugs.add(slug)

                # Опциональные поля
                material_val = row[col_map["material"]] if col_map.get("material") is not None else None
                material = str(material_val).strip() if material_val is not None else None
                if material and material.lower() == "none":
                    material = None

                color_val = row[col_map["color"]] if col_map.get("color") is not None else None
                color = str(color_val).strip() if color_val is not None else None
                if color and color.lower() == "none":
                    color = None

                desc_val = row[col_map["description"]] if col_map.get("description") is not None else None
                description = str(desc_val).strip() if desc_val is not None else None
                if description and description.lower() == "none":
                    description = None

                img_val = row[col_map["image_url"]] if col_map.get("image_url") is not None else None
                image_url = str(img_val).strip() if img_val is not None else None
                if image_url and image_url.lower() == "none":
                    image_url = None

                # Таймер и промо
                show_timer = False
                if "show_timer" in col_map:
                    st_val = str(row[col_map["show_timer"]]).strip().lower()
                    show_timer = st_val in ["true", "1", "yes", "да"]

                from datetime import datetime
                def parse_date(val):
                    if not val or str(val).lower() == "none": return None
                    if isinstance(val, datetime): return val
                    try:
                        return datetime.fromisoformat(str(val))
                    except:
                        return None

                product = Product(
                    name=name,
                    slug=slug,
                    new_price=new_price,
                    old_price=old_price,
                    category_id=category_id,
                    material=material,
                    color=color,
                    description=description,
                    image_url=image_url,
                    is_active=is_active,
                    availability_status=availability_status,
                    show_timer=show_timer,
                    promo_start=parse_date(row[col_map["promo_start"]]) if "promo_start" in col_map else None,
                    promo_end=parse_date(row[col_map["promo_end"]]) if "promo_end" in col_map else None,
                    dimensions=get_val("dimensions"),
                    legs_material=get_val("legs_material"),
                    tabletop_material=get_val("tabletop_material"),
                    tabletop_thickness=get_val("tabletop_thickness"),
                    floor_clearance=get_val("floor_clearance"),
                    max_load=get_val("max_load"),
                    legs_adjustment=get_val("legs_adjustment"),
                    tabletop_color=get_val("tabletop_color"),
                    footings=get_val("footings"),
                    warranty=get_val("warranty"),
                    delivery_format=get_val("delivery_format"),
                    supports=get_val("supports"),
                    country=get_val("country"),
                    series=get_val("series"),
                    characteristics=char_dict,
                )
                db.add(product)
                created += 1

            except Exception as e:
                errors.append(BulkImportError(row=row_idx, error=str(e)))

        if created > 0:
            await db.commit()
        
        logger.info(f"Bulk import finished: {created} products created, {skipped} skipped, {len(errors)} errors")
        wb.close()
        return BulkImportResponse(created=created, skipped=skipped, errors=errors)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка обработки файла: {str(e)}")


# =====================================================================
# ADMIN ENDPOINTS — CATEGORY CRUD
# =====================================================================

@router.post("/categories", response_model=CategoryResponse, status_code=201)
async def create_category(
    category_in: CategoryCreate,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_roles(["admin"]))
):
    db_cat = Category(**category_in.model_dump())
    db.add(db_cat)
    await db.commit()
    await db.refresh(db_cat)
    return db_cat

@router.patch("/categories/{cat_id}", response_model=CategoryResponse)
async def update_category(
    cat_id: int,
    category_in: CategoryUpdate,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_roles(["admin"]))
):
    result = await db.execute(select(Category).where(Category.id == cat_id))
    db_cat = result.scalar_one_or_none()
    if not db_cat:
        raise HTTPException(status_code=404, detail="Категория не найдена")

    update_data = category_in.model_dump(exclude_unset=True)
    for k, v in update_data.items():
        setattr(db_cat, k, v)

    await db.commit()
    await db.refresh(db_cat)
    return db_cat

@router.delete("/categories/{cat_id}")
async def delete_category(
    cat_id: int,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_roles(["admin"]))
):
    result = await db.execute(select(Category).where(Category.id == cat_id))
    db_cat = result.scalar_one_or_none()
    if not db_cat:
        raise HTTPException(status_code=404, detail="Категория не найдена")
    
    # Optional: check if category has products
    prod_result = await db.execute(select(Product).where(Product.category_id == cat_id))
    if prod_result.scalars().first():
         raise HTTPException(status_code=400, detail="Нельзя удалить категорию, в которой есть товары")

    await db.delete(db_cat)
    await db.commit()
    return {"message": "Категория удалена"}
